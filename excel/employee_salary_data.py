import openpyxl
from openpyxl.styles import Font, PatternFill
import random

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employee Salary Data"

# Define grade structure
grades = {
    'A': {'min': 80000, 'max': 120000, 'code': 'SA'},
    'B': {'min': 60000, 'max': 79999, 'code': 'SB'},
    'C': {'min': 40000, 'max': 59999, 'code': 'SC'},
    'D': {'min': 25000, 'max': 39999, 'code': 'SD'},
    'E': {'min': 15000, 'max': 24999, 'code': 'SE'}
}

# Headers
headers = ['Employee ID', 'Employee Name', 'Grade', 'Salary Code', 'Salary']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Generate 1000 employees
for i in range(2, 1002):
    emp_id = f"EMP{i-1:04d}"
    emp_name = f"Employee {i-1}"
    
    # Random grade selection
    grade = random.choice(list(grades.keys()))
    salary_code = grades[grade]['code']
    salary = random.randint(grades[grade]['min'], grades[grade]['max'])
    
    ws.cell(row=i, column=1, value=emp_id)
    ws.cell(row=i, column=2, value=emp_name)
    ws.cell(row=i, column=3, value=grade)
    ws.cell(row=i, column=4, value=salary_code)
    ws.cell(row=i, column=5, value=salary)

# Auto-adjust column widths
for column in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_length + 2

# Save the file
wb.save("Employee_Salary_Data.xlsx")
print("Excel file 'Employee_Salary_Data.xlsx' created successfully with 1000 employees!")